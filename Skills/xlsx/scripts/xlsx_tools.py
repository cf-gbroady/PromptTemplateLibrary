#!/usr/bin/env python3
"""
xlsx_tools.py

Helper utilities for xlsx skill workflows in Code Interpreter or local automation.

Capabilities:
- Recalculate workbooks with LibreOffice when available.
- Scan workbooks for blocking Excel formula errors.
- Add or update a Sources worksheet for hardcoded source documentation.
- Set workbook calculation mode to automatic.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BLOCKING_FORMULA_ERRORS = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#N/A",
    "#NAME?",
    "#NUM!",
    "#NULL!",
}

SOURCE_HEADERS = [
    "Source ID",
    "Workbook Location",
    "Value / Assumption",
    "Source Type",
    "Source Name",
    "Source Date",
    "Specific Reference",
    "URL / Path",
    "Notes",
]


@dataclass
class FormulaError:
    sheet: str
    cell: str
    value: str


@dataclass
class ValidationResult:
    path: str
    formula_error_count: int
    formula_errors: list[FormulaError]
    workbook_calculation_mode: Optional[str]

    @property
    def ok(self) -> bool:
        return self.formula_error_count == 0


def find_libreoffice() -> Optional[str]:
    for name in ("soffice", "libreoffice"):
        found = shutil.which(name)
        if found:
            return found
    return None


def set_automatic_calculation(path: str | Path) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False)
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        try:
            calc.calcMode = "auto"
            calc.fullCalcOnLoad = True
            calc.forceFullCalc = True
        except Exception:
            pass
    wb.save(workbook_path)


def recalc_workbook(path: str | Path, timeout: int = 60) -> Path:
    source = Path(path).resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    office = find_libreoffice()
    if not office:
        raise RuntimeError("LibreOffice/soffice is not available in this environment.")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        working = tmpdir_path / source.name
        shutil.copy2(source, working)
        cmd = [
            office,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(tmpdir_path),
            str(working),
        ]
        completed = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=timeout,
        )
        if completed.returncode != 0:
            raise RuntimeError(
                "LibreOffice recalculation failed: "
                + (completed.stderr.strip() or completed.stdout.strip())
            )
        recalculated = tmpdir_path / source.name
        if not recalculated.exists():
            raise RuntimeError("LibreOffice did not produce a recalculated workbook.")
        shutil.copy2(recalculated, source)
    return source


def scan_formula_errors(path: str | Path) -> list[FormulaError]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False)
    errors: list[FormulaError] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() in BLOCKING_FORMULA_ERRORS:
                    errors.append(FormulaError(ws.title, cell.coordinate, value.strip()))
    return errors


def validate_workbook(path: str | Path) -> ValidationResult:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    calc_mode = None
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc_mode = getattr(calc, "calcMode", None)
    wb.close()
    errors = scan_formula_errors(workbook_path)
    return ValidationResult(
        path=str(workbook_path),
        formula_error_count=len(errors),
        formula_errors=errors,
        workbook_calculation_mode=calc_mode,
    )


def recalc_and_validate(path: str | Path, timeout: int = 60) -> ValidationResult:
    set_automatic_calculation(path)
    recalc_workbook(path, timeout=timeout)
    return validate_workbook(path)


def ensure_sources_sheet(path: str | Path, rows: Optional[list[dict[str, str]]] = None) -> None:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb["Sources"] if "Sources" in wb.sheetnames else wb.create_sheet("Sources")

    for col_idx, header in enumerate(SOURCE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    if rows:
        start_row = max(ws.max_row + 1, 2)
        for row_offset, source_row in enumerate(rows):
            for col_idx, header in enumerate(SOURCE_HEADERS, start=1):
                ws.cell(row=start_row + row_offset, column=col_idx, value=source_row.get(header, ""))

    widths = {"A": 12, "B": 24, "C": 32, "D": 18, "E": 28, "F": 14, "G": 36, "H": 36, "I": 36}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(workbook_path)


def load_source_rows(path: str | Path) -> list[dict[str, str]]:
    source_path = Path(path)
    if source_path.suffix.lower() == ".json":
        data = json.loads(source_path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            raise ValueError("JSON source rows must be a list of objects.")
        return [{str(k): "" if v is None else str(v) for k, v in row.items()} for row in data]
    if source_path.suffix.lower() == ".csv":
        with source_path.open("r", encoding="utf-8-sig", newline="") as f:
            return [dict(row) for row in csv.DictReader(f)]
    raise ValueError("Source rows must be provided as .json or .csv")


def result_to_dict(result: ValidationResult) -> dict:
    data = asdict(result)
    data["ok"] = result.ok
    return data


def print_validation_result(result: ValidationResult) -> None:
    print(json.dumps(result_to_dict(result), indent=2))


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="xlsx workbook helper utilities")
    sub = parser.add_subparsers(dest="command", required=True)

    p_recalc = sub.add_parser("recalc")
    p_recalc.add_argument("workbook")
    p_recalc.add_argument("--timeout", type=int, default=60)

    p_validate = sub.add_parser("validate")
    p_validate.add_argument("workbook")

    p_rv = sub.add_parser("recalc-and-validate")
    p_rv.add_argument("workbook")
    p_rv.add_argument("--timeout", type=int, default=60)

    p_sources = sub.add_parser("ensure-sources")
    p_sources.add_argument("workbook")
    p_sources.add_argument("--rows", help="Optional .json or .csv source rows file")

    args = parser.parse_args(list(argv) if argv is not None else None)

    if args.command == "recalc":
        set_automatic_calculation(args.workbook)
        recalc_workbook(args.workbook, timeout=args.timeout)
        return 0
    if args.command == "validate":
        result = validate_workbook(args.workbook)
        print_validation_result(result)
        return 0 if result.ok else 1
    if args.command == "recalc-and-validate":
        result = recalc_and_validate(args.workbook, timeout=args.timeout)
        print_validation_result(result)
        return 0 if result.ok else 1
    if args.command == "ensure-sources":
        rows = load_source_rows(args.rows) if args.rows else None
        ensure_sources_sheet(args.workbook, rows=rows)
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
